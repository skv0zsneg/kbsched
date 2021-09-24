from os import path
from typing import Any
from typing import Union
from requests import get
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from collections import defaultdict
from openpyxl.utils import get_column_letter


KEY_WORD_INSTITUTES = ('КБиСП',)
FILE_FORMAT = '.xlsx'
FILE_NAME_FOR_CURRENT_SCHEDULE = 'schedule' + FILE_FORMAT
CURRENT_PATH = path.split(__file__)[0]


class Parse:
    def __init__(self, url: str):
        self.bs = BeautifulSoup(get(url).text, 'html.parser')

    def get_links_for_courses(self) -> defaultdict:
        course_links = defaultdict(list)
        for bs_link in self.bs.find_all('a', attrs={'class': 'uk-link-toggle'}):
            href = bs_link.get('href')
            if any(map(lambda w: w in href, KEY_WORD_INSTITUTES)) and FILE_FORMAT in href:
                course_number = bs_link.find('div', attrs={'class': 'uk-link-heading uk-margin-small-top'})
                course_links[course_number.text.strip()].append(href)
        return course_links


class GetSchedule:
    def __init__(self, url: str, kb_model: Any = None):
        self.url = url
        self.kb_model = kb_model
        self.parse = Parse(self.url) 

    def get_links(self):
        course_links = self.parse.get_links_for_courses()
        for course_name, links in course_links.items():
            for link in links:
                self.kb_model.db.execute(
                    self.kb_model.schedule_links_table.insert_into('NULL', course_name, link)
                )

    def parse_all(self):
        for link in self.kb_model.get_elements('schedule_link', table=self.kb_model.schedule_links_table):
            print(link)

    def parse_all_for_course(self, course: Union[int, str]):
        course = int(course[0]) if isinstance(course, str) else course
        for link in self.kb_model.get_elements('schedule_link',
                                               table=self.kb_model.schedule_links_table,
                                               where=f"course_name == '{str(course)} курс'")[0]:
            with open(path.join(CURRENT_PATH, f'{FILE_NAME_FOR_CURRENT_SCHEDULE}'), 'wb') as f:
                f.write(get(link).content)
            wb = load_workbook(FILE_NAME_FOR_CURRENT_SCHEDULE)
            list_1 = wb.active
            print(list_1['F5'].value)
